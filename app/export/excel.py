import os
import tempfile
from typing import Dict, Any, List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from ..rdl.expression import evaluate_expression, ExpressionContext


def export_to_excel(report_def: Dict, data: Dict[str, Any], report_name: str) -> str:
    """
    Export report data to Excel file.

    Returns: Path to the generated Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = report_name[:31]  # Excel sheet name limit

    tablixes = report_def.get('tablixes', [])
    if not tablixes:
        ws.cell(row=1, column=1, value="No data")
        temp_path = os.path.join(tempfile.gettempdir(), f"{report_name}.xlsx")
        wb.save(temp_path)
        return temp_path

    tablix = tablixes[0]
    columns = tablix.get('columns', [])
    rows = data.get('primary', [])
    lookup_tables = data.get('lookup_tables', {})

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="207176", end_color="207176", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write header row
    for col_idx, col in enumerate(columns, 1):
        header_text = col.get('header_text', '')
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data rows
    for row_idx, row in enumerate(rows, 2):
        context = ExpressionContext(row, row_idx - 1, lookup_tables)

        for col_idx, col in enumerate(columns, 1):
            expression = col.get('field_expression', '')
            value = evaluate_expression(expression, context)

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    # Auto-fit column widths
    for col_idx, col in enumerate(columns, 1):
        column_letter = get_column_letter(col_idx)
        max_length = len(col.get('header_text', ''))

        for row_idx in range(2, len(rows) + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))

        adjusted_width = min(max_length + 2, 50)  # Cap at 50
        ws.column_dimensions[column_letter].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Save to temp file
    temp_path = os.path.join(tempfile.gettempdir(), f"{report_name}.xlsx")
    wb.save(temp_path)

    return temp_path
